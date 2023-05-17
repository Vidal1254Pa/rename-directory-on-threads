import openpyxl as xlxs
import threading as th
import numpy as np
import os
import shutil
from time import sleep
excel_instace = xlxs.load_workbook('./nameDirs.xlsx')
data_frame = excel_instace.active
DIRECORY_FOTOS = './FOTOS/'
'''
iterate over all rows in specific column
'''
data = []
for row in range(1, data_frame.max_row):
    _row = [row,]
    for col in data_frame.iter_cols(1, data_frame.max_column):
        _row.append(col[row].value)
    data.append(_row)
'''
numero de hilos
'''
counter_data_distribution = data_frame.max_row-1
counter_data_capture = 0
task_hilos = [
    [], [], [], [], [], [], [], [], [], []
]
instance_block_data = np.array(data)


def defier_ranger(hilos):
    global counter_data_distribution
    global counter_data_capture
    resto = counter_data_capture-counter_data_distribution
    resto = abs(resto)
    range_number = []
    for i in range(hilos+1):
        if counter_data_capture == counter_data_distribution:
            print('cumplio todo')
            break
        resto -= 1
        counter_data_capture += 1
        range_number.append(resto)
    aux=list(reversed(range_number))
    print(aux)
    return list(reversed(range_number))


def assing_task(index=None):
    global hilos
    global task_hilos
    global instance_block_data
    for i in range(hilos):
        range_hilo = defier_ranger(hilos)
        data_task = np.take(instance_block_data, range_hilo, axis=0)
        task_hilos[index or i] = (data_task)


def assing_new_task(index):
    global hilos
    global task_hilos
    global instance_block_data
    range_hilo = defier_ranger(hilos)
    data_task = np.take(instance_block_data, range_hilo, axis=0)
    task_hilos[index] = data_task


def resolve_task(array_task):
    print('hola ' + str(th.currentThread().getName()))
    for data in array_task:
        index, init, target,code = data
        # verifier exist directory
        if os.path.exists(DIRECORY_FOTOS+str(init)) and os.path.exists(DIRECORY_FOTOS+str(target)+'-'+str(code)):
            shutil.rmtree(DIRECORY_FOTOS+str(init))
            continue
        if os.path.exists(DIRECORY_FOTOS+str(init)):
            os.rename(DIRECORY_FOTOS+str(init),
                      DIRECORY_FOTOS+str(target)+'-'+str(code))
            print('renombrado de '+str(code)+' a '+str(target)+'-'+str(code))


hilos = 10
'''primara asignacion de tareas'''
assing_task()
divider_task = data_frame.max_row//hilos
thread_instance = []
for i in range(hilos):
    thread_instance.append(
        th.Thread(target=resolve_task, args=(task_hilos[i],), name='hilo'+str(i)))

for i in range(hilos):
    thread_instance[i].start()

while True:
    if counter_data_capture == counter_data_distribution:
        print('termino')
        break
    sleep(1)
    try:
        for i in range(hilos):
            if not thread_instance[i].is_alive():
                thread_instance[i].join()
                assing_new_task(i)
                thread_instance[i] = th.Thread(
                    target=resolve_task, args=(task_hilos[i],), name='hilo'+str(i))
                thread_instance[i].start()
    except Exception as e:
        pass
    finally:
        print(str(counter_data_capture)+' de '+str(counter_data_distribution))
        pass
